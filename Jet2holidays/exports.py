import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill
from pymongo import MongoClient
import pandas as pd

# ------------------------------
# Step 1: Setup
# ------------------------------
today = datetime.datetime.now().strftime("%d-%m-%Y")
mongo_uri = "mongodb://localhost:27017"  # Change if needed
client = MongoClient(mongo_uri)
db = client["PriceMate_alibaba"]
collection = db[f"Product_Data_2026_02_09"]  # Update collection name

# ------------------------------
# Step 2: Fetch data from MongoDB
# ------------------------------
cursor = collection.find()
data = list(cursor)

# Remove MongoDB’s internal _id field if not needed
for item in data:
    item.pop('_id', None)

# Convert to DataFrame
df = pd.DataFrame(data)

# ------------------------------
# Step 3: Define your header order (series)
# ------------------------------
header_order = [   # jet2holidays header 

"Crawled Date"
"Website"
"Check-In Date"
"Check-Out Date"
"Nights"
"Pax"
"Destination type"
"Website Hotel Name"
"Website Hotel ID"
"Matched Hotel Name"
"Matched Hotel ID"
"Website Rating"
"City"
"Country"
"Source Market"
"Room Name"
"Board Type"
"Total Price"
"Currency"
"Flight Status"
"Departure Airport"
"Arrival Airport"
"Out Airline"
"Out Flight code"
"Out Depart Date"
"Out Arr Date"
"In Airline"
"In Flight code"
"In Depart Date"
"In Arr Date"
"Baggage"
"Transfer"
]
# header_order = [
# "name"
# "groupId"
# "gtin"
# "productNo"
# "price"
# "pricePerUnit"
# "netQuantity"
# "manufacturer"
# "brand"
# "promotionText"
# "rank"
# "imageURL"
# "outOfStock"
# "sponsore"
# "breadcrumb"
# "badge"
# "url"
# "avgRating"
# "commentCount"
# "CategoryJourney"
# "originalURL"
# "Pagetype"
# "PageNo"
# "postcode"
# "store Name"
# "store Address"
# "store ID"
# "meta_info"
# "file_name"
# "crawl_time"
# ]

# Keep only columns that exist in DataFrame
ordered_cols = [col for col in header_order if col in df.columns]

# Optional: append any extra columns not in your series at the end
extra_cols = [col for col in df.columns if col not in ordered_cols]
df = df[ordered_cols + extra_cols]

# ------------------------------
# Step 4: Export to Excel
# ------------------------------
output_file = f"XWIZ_Jet2Holidays_Sample_{today}.xlsx"
with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
    writer.book.strings_to_urls = False
    df.to_excel(writer, index=False)

# ------------------------------
# Step 5: Load workbook for formatting
# ------------------------------
wb = load_workbook(output_file)
ws = wb.active

# Thin border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Light grey fill for header row
header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# Adjust column width & apply styles
for col in ws.columns:
    col_letter = col[0].column_letter
    ws.column_dimensions[col_letter].width = 25  # Adjust as needed
    for cell in col:
        cell.border = thin_border
        if cell.row == 1:  # Header row
            cell.fill = header_fill

# ------------------------------
# Step 6: Save workbook
# ------------------------------
wb.save(output_file)

print(f"Data exported to '{output_file}' with headers in series, borders, and grey headers.")
